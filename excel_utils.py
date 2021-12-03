"""Collection of helper functions for working with Excel Spreadsheets."""
import openpyxl
import pandas as pd


def write_excel(df: pd.DataFrame, filename: str, filter_sheet: bool = True) -> None:
    """Write a Pandas DataFrame to an Excel Spreadsheet

    Writes a Pandas DataFrame to an Excel Spreadsheet and formats the spreadsheet.
    The formatting applied is to adust the column widths and turn on filtering.

    Args:
        df: The Pandas DataFrame to export as Excel.
        filename: str of the name of the file to create.
        filter_sheet: bool to indicate turning on Excel filtering. Default: True

    Raises:
        OSError: Raised if the there are issues creating/saving the spreadsheet
    """
    try:
        df.to_excel(filename, index=False, engine='openpyxl')
    except OSError as e:
        raise e

    try:
        workbook = openpyxl.load_workbook(filename)
    except OSError as e:
        raise e

    if filter_sheet:
        filter_workbook(workbook)

    adjust_column_widths(workbook)

    try:
        workbook.save(filename)
        workbook.close()
    except OSError as e:
        raise e


def filter_workbook(workbook: openpyxl.Workbook) -> None:
    """Turns on data filtering for columns in the spreadsheet

    filter_workbook takes an openpyxl Workbook and iterates through the
    sheets and enables the data filter for the columns in the sheets.

    Args:
        workbook: the openpyxl Workbook to turn on data filtering.

    Returns:
        None. The workbook is modified in place.

    Raises:
        None.
    """
    for sheetname in workbook.sheetnames:
        sheet = workbook[sheetname]
        sheet.auto_filter.ref = sheet.dimensions


def adjust_column_widths(workbook: openpyxl.Workbook, magic: int = 5) -> None:
    """Adjusts the column widths to account for the data in the column's cells

    adjust_column_widths takes an openpyxl Workbook and changes the widths of
    the columns to match the size of the longest cell in that column.

    Args:
        workbook: the openpyxl Workbook to adjust columns for.
        magic: Magic number to slightly increase the column size. Default: 5
          This accounts for the max length of a cell not quite being a large
          enough value to look correct when viewed in Excel.

    Returns:
        None. The Workbook is modified in place.

    Raises:
        None.
    """

    for sheetname in workbook.sheetnames:
        sheet = workbook[sheetname]
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells) + magic
            sheet.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = length
